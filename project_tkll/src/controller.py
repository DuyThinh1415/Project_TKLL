#!/usr/bin/env python
import rospy
from demo_pakage.msg import Num
import openpyxl 
  
#----------------------------SETTING DEFAULT---------------------------#
# CHANGE LANE - DEFAULT VALUE OF OPTINAL VALUE

DEFAULT_CHANGE_LANE_ALPHA = 30
DEFAULT_CHANGE_LANE_V_ANGULAR = 0.03
DEFAULT_CHANGE_LANE_DELTA_D = 0.5
DEFAULT_CHANGE_LANE_NO_USE = -1

# DETECT TRAFFIC SIGNAL DEFAULT VALUE
DEFAULT_DETECT_SIGNAL_LIVE_TIME = 15

# INVALID COMMAND
IVL_CMD = "Invalid command"

# EXCEL CONSTANT
COL_STT = 1
COL_HEADER = 2
ROW_OF_NAME_COLS = 1

#----------------------------INSTALL LIBRARY---------------------------#

# Give the location of the file 
path = "/home/thinh/ROS/demo/src/demo_pakage/src/cmd.xlsx"

work_book_obj = openpyxl.load_workbook(path)
work_book_active_obj = work_book_obj.active
num_row = work_book_active_obj.max_row
num_col = work_book_active_obj.max_column
  
#----------------------------FUNCTION---------------------------#
def solve(cmd_split_arr, row_of_detected_header):
    resultArray = []
    index = 1
    for col in range(COL_HEADER + 1, num_col + 1):
        if(work_book_active_obj.cell(row=row_of_detected_header, column=col).value == None):
            resultArray.append(DEFAULT_CHANGE_LANE_NO_USE)        
        else:
            if(cmd_split_arr[index] != '-'):
                resultArray.append(float(cmd_split_arr[index]))                
            else:
                resultArray.append(float(work_book_active_obj.cell(row=row_of_detected_header, column=col).value))
            index = index + 1
                
    response_header = work_book_active_obj.cell(row=row_of_detected_header, column=COL_STT).value
    response_base_arg = int(resultArray[0])
    respose_int_1 = int(resultArray[1])
    respose_int_2 = int(resultArray[2])
    respose_int_3 = int(resultArray[3])
    respose_float_1 = float(resultArray[4])
    respose_float_2 = float(resultArray[5])
    respose_float_3 = float(resultArray[6])
    return [response_header, response_base_arg, respose_int_1, respose_int_2, respose_int_3, respose_float_1, respose_float_2, respose_float_3]

pub = rospy.Publisher('controller_topic', Num, queue_size=10)
rospy.init_node('Controller', anonymous=True)

while not rospy.is_shutdown():
    cmd = input("~ ")
    splitCmd = cmd.split()        
    
    # Check Header of command
    if(len(splitCmd) >= 1):
        header = splitCmd[0]
        isExistHeader = False
        for row in range(ROW_OF_NAME_COLS + 1, num_row + 1):
            if(work_book_active_obj.cell(row=row, column=COL_HEADER).value == header):
                isExistHeader = True                
                break                
                
        if(isExistHeader):
            returnData = solve(splitCmd, row)
            msg = Num()
            msg.header = returnData[0]
            msg.base_arg = returnData[1]
            msg.int_1 = returnData[2]
            msg.int_2 = returnData[3]
            msg.int_3 = returnData[4]
            msg.float_1 = returnData[5]
            msg.float_2 = returnData[6]
            msg.float_3 = returnData[7]
            pub.publish(msg)
        else:        
            print(IVL_CMD)
    else:
        print(IVL_CMD)

    rate = rospy.Rate(10) # 10hz
    rate.sleep()